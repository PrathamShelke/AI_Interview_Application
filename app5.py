import streamlit as st
import openai
import pandas as pd
from io import BytesIO
from audio_recorder_streamlit import audio_recorder
import os
from openpyxl import load_workbook

# Set your OpenAI API key
os.environ["OPENAI_API_KEY"] = "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"
client = openai.OpenAI()

def text_to_speech(text):
    response = client.audio.speech.create(
        model="tts-1",
        voice="alloy",
        input=text
    )
    audio_data = BytesIO(response.content)
    return audio_data

def transcribe_audio(audio_bytes):
    transcript_response = client.audio.translations.create(
        model="whisper-1",
        file=audio_bytes,
        response_format="text"
    )
    return transcript_response

def save_to_excel(question, answer, filename="responses.xlsx"):
    df = pd.DataFrame({'Question': [question], 'Answer': [answer]})
    if not os.path.isfile(filename):
        df.to_excel(filename, index=False)
    else:
        book = load_workbook(filename)
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book

        if 'Sheet1' in book.sheetnames:
            startrow = book['Sheet1'].max_row

            df.to_excel(writer, index=False, header=False, startrow=startrow)
        else:
            df.to_excel(writer, index=False, header=True)

        writer.save()

questions = [
    "What is your favorite color?",
    "How do you spend your free time?"
]

st.title("Q&A Session with ChatGPT")

if 'current_question_index' not in st.session_state:
    st.session_state['current_question_index'] = 0
    st.session_state['waiting_for_answer'] = False

if st.session_state['current_question_index'] < len(questions):
    if not st.session_state['waiting_for_answer']:
        question = questions[st.session_state['current_question_index']]
        st.write("Question:")
        st.write(question)

        tts_audio = text_to_speech(question)
        st.audio(tts_audio, format="audio/mp3")
        st.session_state['waiting_for_answer'] = True

    audio_bytes = audio_recorder()

    if audio_bytes and st.button("Submit Response"):
        answer_text = transcribe_audio(BytesIO(audio_bytes))
        st.write(f"Your response: {answer_text}")

        save_to_excel(question, answer_text)

        st.session_state['current_question_index'] += 1
        st.session_state['waiting_for_answer'] = False

if st.session_state['current_question_index'] >= len(questions):
    st.write("All questions have been answered.")
